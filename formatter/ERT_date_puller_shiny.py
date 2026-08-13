from datetime import datetime
import os
import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from shiny import App, reactive, render, req, ui

###########################################################################################
# Input colors used by the parser
GREEN = "#FFE2EFDA"  # Entry to the GOT or TRAY
RED = "#FFFFCCCC"    # Departure from the GOT (disappearance, death, transfer)
BLUE = "#FF4D93D9"   # Still Alive as of 12/31/24


# Helper functions

def normalize_color(color):
    """Normalize color values to an 8-digit ARGB hex string."""
    if color is None:
        return None

    if isinstance(color, str):
        value = color.strip().upper()
        return value if value.startswith("#") else f"#{value}"

    if hasattr(color, "rgb") and color.rgb:
        value = str(color.rgb).strip().upper()
        return value if value.startswith("#") else f"#{value}"

    if hasattr(color, "value") and color.value:
        value = str(color.value).strip().upper()
        return value if value.startswith("#") else f"#{value}"

    if hasattr(color, "type") and getattr(color, "type") == "rgb":
        value = str(getattr(color, "value", "")).strip().upper()
        return value if value.startswith("#") else f"#{value}"

    return str(color).upper()


def parse_date(value):
    """Convert an Excel cell into a datetime object when possible."""
    if value is None:
        return None

    if isinstance(value, datetime):
        return value

    try:
        return datetime.strptime(str(value), "%m/%d/%Y")
    except ValueError:
        return None


def get_cell_date(value, color, start_date=None, is_intro=False):
    """Resolve a cell date, including census handling for green and red cells."""
    if value is None:
        return None

    if isinstance(value, datetime):
        return value

    parsed = parse_date(value)
    if parsed is not None:
        return parsed

    census_match = re.search(r"(\d{4})\s*census|census\s*(\d{4})", str(value or ""), flags=re.I)
    if census_match:
        year = int(census_match.group(1) or census_match.group(2))

        if is_intro:
            return datetime(year, 12, 31)

        if color == RED:
            if start_date is not None and start_date.year == year:
                end_of_year = datetime(year, 12, 31)
                return start_date + (end_of_year - start_date) / 2

            return datetime(year, 7, 1)

    return None


def calculate_yearfrac(start_date, stop_date):
    """Calculate the time in years between two dates."""
    if start_date is None or stop_date is None:
        return None

    if stop_date < start_date:
        start_date, stop_date = stop_date, start_date

    days = (stop_date - start_date).days
    return days / 365.25


def sheet_prefix(sheet_name):
    """Build a compact prefix from a worksheet name."""
    words = re.findall(r"[A-Za-z0-9]+", sheet_name)
    return "".join(word[0].upper() for word in words)


def parse_ert_workbook(workbook_path, sheet_index_path):
    """Parse an ERT workbook into a table of intro/exit records."""
    config_df = pd.read_excel(sheet_index_path)
    names_index = dict(zip(config_df["Species"], config_df["Start Row"]))

    wb = load_workbook(workbook_path)
    records = []

    for ws in wb.worksheets:
        sheet_name = ws.title
        if sheet_name not in names_index:
            continue

        prefix = sheet_prefix(sheet_name)
        start_row = names_index[sheet_name]
        sheet_person_num = 0

        for col in ws.iter_cols():
            intro_date = None
            stop_date = None
            stop_status = None
            had_red_census = False
            had_green_census = False

            for cell in col[start_row - 1:]:
                cell_value = cell.value
                is_blank = cell_value is None or (isinstance(cell_value, str) and not str(cell_value).strip())

                fill = cell.fill
                color = None
                if fill is not None:
                    color = normalize_color(getattr(fill, "fgColor", None))
                    if color is None:
                        color = normalize_color(getattr(fill, "start_color", None))

                is_red = color == RED
                is_green = color == GREEN
                is_blue = color == BLUE
                has_fill_color = is_red or is_green or is_blue

                if is_blank and not has_fill_color:
                    continue

                is_event = has_fill_color or isinstance(cell_value, (datetime, str)) or (
                    isinstance(cell_value, str) and bool(str(cell_value).strip())
                )
                if not is_event:
                    continue

                is_census = bool(re.search(r"(\d{4})\s*census|census\s*(\d{4})", str(cell_value or ""), flags=re.I))
                if is_census and is_red:
                    had_red_census = True
                if is_census and is_green:
                    had_green_census = True

                if is_blue is True:
                    stop_date = datetime(1800, 1, 1)
                    stop_status = "still_alive"

                date = get_cell_date(cell_value, color, intro_date, is_intro=intro_date is None)
                if date is None:
                    continue

                if intro_date is None:
                    intro_date = date
                    continue

                if (is_red or is_blue) and stop_date is None:
                    stop_date = date
                    stop_status = "1" if is_blue else "0"
                    break

            if intro_date is not None and stop_date is not None:
                sheet_person_num += 1
                entity_id = f"{prefix}{sheet_person_num}"

                duration_days = (stop_date - intro_date).days
                yearfrac = calculate_yearfrac(intro_date, stop_date)

                records.append(
                    {
                        "Species": sheet_name,
                        "Individual": entity_id,
                        "Intro_date": intro_date,
                        "Exit_date": stop_date,
                        "duration_days": duration_days,
                        "yearfrac": yearfrac,
                        "status": stop_status
                        #"red_census": "Yes" if had_red_census else "No",
                        #"green_census": "Yes" if had_green_census else "No",
                    }
                )

    expected_columns = [
        "Species",
        "Individual",
        "Intro_date",
        "Exit_date",
        "duration_days",
        "yearfrac",
        "status"
    ]

    df = pd.DataFrame(records, columns=expected_columns)
    if not df.empty:
        df = df.sort_values(by=["Species", "Intro_date"]).reset_index(drop=True)
    else:
        df = df.reset_index(drop=True)

    return df


def _get_uploaded_path(file_info):
    """Normalize Shiny file input into a filesystem path string."""
    if file_info is None:
        return None
    if isinstance(file_info, list):
        file_info = file_info[0] if file_info else None
    if file_info is None:
        return None
    return getattr(file_info, "datapath", None)


def save_results_to_excel(df, output_path):
    """Save the parsed output DataFrame to an Excel workbook when a path is provided."""
    if df is None:
        return None

    output_path_value = str(output_path or "").strip()
    if not output_path_value:
        return None

    output_path_obj = Path(output_path_value)
    if output_path_obj.suffix.lower() != ".xlsx":
        output_path_obj = output_path_obj.with_suffix(".xlsx")

    output_path_obj.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(output_path_obj, index=False)
    return str(output_path_obj)


app_ui = ui.page_fluid(
    ui.h2("ERT Parser in Shiny for Python"),
    ui.p("Upload the main ERT workbook and the sheet-index workbook, then run the parser."),
    ui.row(
        ui.column(
            6,
            ui.input_file("workbook", "ERT workbook", accept=[".xlsx"], multiple=False),
        ),
        ui.column(
            6,
            ui.input_file("index_file", "Sheet index workbook", accept=[".xlsx"], multiple=False),
        ),
    ),
    ui.input_text("output_path", "Save parsed workbook to (optional)", value=""),
    ui.p("Enter a full file path such as C:/temp/parsed_output.xlsx. If left blank, the parsed output is only shown in the table.", style="font-size: 12px; color: gray;"),
    ui.input_action_button("run_parser", "Run parser", class_="btn-primary"),
    ui.output_text_verbatim("save_status"),
    ui.output_table("results_table"),
)


def server(input, output, session):
    save_status_msg = reactive.Value("")

    @reactive.event(input.run_parser)
    def parsed_df():
        workbook_path = _get_uploaded_path(input.workbook())
        index_path = _get_uploaded_path(input.index_file())
        req(workbook_path, index_path)

        df = parse_ert_workbook(workbook_path, index_path)
        save_path = input.output_path()
        if save_path:
            try:
                saved_path = save_results_to_excel(df, save_path)
                if saved_path:
                    save_status_msg.set(f"Saved parsed workbook to {saved_path}")
                else:
                    save_status_msg.set("No save path provided; workbook was only displayed.")
            except Exception as exc:
                save_status_msg.set(f"Unable to save workbook: {exc}")
        else:
            save_status_msg.set("No save path provided; workbook was only displayed.")

        return df

    @render.text
    def save_status():
        return save_status_msg()

    @render.table
    def results_table():
        df = parsed_df()
        if df is None:
            return pd.DataFrame({"Status": ["Upload both files and click Run parser."]})
        return df


app = App(ui=app_ui, server=server)


if __name__ == "__main__":
    app.run()
