#Loads modules
from openpyxl import load_workbook
from datetime import datetime
import pandas as pd
import re
import math


#Inputs
#FILE_PATH = r"C:\Users\jgasink\Desktop\ERT Project\Working Copies\Species_ERT_Workbook_7-1-26.xlsx"
FILE_PATH = r"C:\Users\jgasink\Desktop\ERT Project\Python\Demo_data.xlsx"

START_ROW = 26


GREEN = "#FFE2EFDA" #Entry to the GOT or TRAY
RED = "#FFFFCCCC" #Departure from the GOT (disappearance, death, transfer)
BLUE = "#FF4D93D9" #Still Alive as of 12/31/24

OUTPUT_FILE = r"C:\Users\jgasink\Desktop\ERT Project\Python\ERT_Translated_Porgy.xlsx"

#Optional allowlist of sheet names to process. Leave empty to process all sheets.
SHEETS_TO_PROCESS = ["Jolthead Porgy", "Saucereye Porgy", "sheepshead porgy"]


#Helper function
#Normalizes color to 8 digit ARGB hex code (e.g. #AARRGGBB)
def normalize_color(color):

    if color is None:
        return None

    if isinstance(color, str):
        value = color.strip().upper()
        return value if value.startswith("#") else f"#{value}"

    if hasattr(color, "rgb") and color.rgb:
        value = str(color.rgb).strip().upper()
        return value if value.startswith("#") else f"#{value}"

    if hasattr(color, "value") and color.value:
        value = str(color.value).strip().upper()
        return value if value.startswith("#") else f"#{value}"

    if hasattr(color, "type") and getattr(color, "type") == "rgb":
        value = str(getattr(color, "value", "")).strip().upper()
        return value if value.startswith("#") else f"#{value}"

    return str(color).upper()

#Converts an Excel cell into a datetime object
def parse_date(value):

    if value is None:
        return None

    if isinstance(value, datetime):
        return value

    try:
        return datetime.strptime(str(value), "%m/%d/%Y")
    except ValueError:
        return None

#Resolves a cell date, including Census year handling for green and red cells.
def get_cell_date(value, color, start_date=None, is_intro=False):

    if value is None:
        return None

    if isinstance(value, datetime):
        return value

    parsed = parse_date(value)
    if parsed is not None:
        return parsed

    census_match = re.search(r"(\d{4})\s*census|census\s*(\d{4})", str(value or ""), flags=re.I)

    if census_match:
        year = int(census_match.group(1) or census_match.group(2))

        if is_intro:
            return datetime(year, 12, 31)

        if color == RED:
            if start_date is not None and start_date.year == year:
                end_of_year = datetime(year, 12, 31)
                return start_date + (end_of_year - start_date) / 2

            return datetime(year, 7, 1)

    return None

#Function that calculates the intro and exit date of an animal into the GOT and gives the yearfrac (Length of time in tank)
def calculate_yearfrac(start_date, stop_date):

    if start_date is None or stop_date is None:
        return None

    if stop_date < start_date:
        start_date, stop_date = stop_date, start_date

    days = (stop_date - start_date).days
    return days / 365.25

#Only runs specified sheets
def should_process_sheet(sheet_name):

    if not SHEETS_TO_PROCESS:
        return True

    allowed_sheets = {name.strip().casefold() for name in SHEETS_TO_PROCESS if name and name.strip()}
    return sheet_name.strip().casefold() in allowed_sheets

#Pulls the first letter of each word in a sheet name and returns it as a string (Collects Species Name)
def sheet_prefix(sheet_name):


    words = re.findall(r"[A-Za-z0-9]+", sheet_name)

    return "".join(word[0].upper() for word in words)


wb = load_workbook(FILE_PATH)

records = []

#Loads the workbook and iterates through each worksheet, processing each column to extract start and stop dates based on cell fill colors. 
#It collects the data into a list of records, which is then converted into a DataFrame and saved to an Excel file.
for ws in wb.worksheets:

    sheet_name = ws.title

    if not should_process_sheet(sheet_name):
        print(f"Skipping {sheet_name}")
        continue

    prefix = sheet_prefix(sheet_name)

    print(f"Processing {sheet_name}")

    sheet_person_num = 0

    for col in ws.iter_cols():

        intro_date = None
        stop_date = None
        stop_status = None
        pending_records = []
        had_red_census = False
        had_green_census = False

        #Scan each column row by row and treat the first meaningful date/census as the introduction,
        #then use the first later red/blue/census entry as the exit.
        for cell in col[START_ROW - 1:]:

            cell_value = cell.value
            is_blank = cell_value is None or (isinstance(cell_value,str) and not str(cell_value).strip())

            fill = cell.fill
            color = None

            if fill is not None:
                color = normalize_color(getattr(fill, "fgColor", None))
                if color is None:
                    color = normalize_color(getattr(fill, "start_color", None))

            is_red = color == RED
            is_green = color == GREEN
            is_blue = color == BLUE
            has_fill_color = is_red or is_green or is_blue 
            
            if is_blank and not has_fill_color:
                continue

            is_event = has_fill_color or isinstance(cell_value, (datetime, str)) or (isinstance(cell_value, str) and bool(str(cell_value).strip()))


            if not is_event:
                continue

            is_census = bool(re.search(r"(\d{4})\s*census|census\s*(\d{4})", str(cell_value or ""), flags=re.I))
            if is_census and is_red:
                had_red_census = True
            elif is_census and is_green:
                had_green_census = True

            date = get_cell_date(cell_value, color, intro_date, is_intro=intro_date is None)
            if date is None:
                continue

            if intro_date is None:
                intro_date = date
                continue

            if is_blue is not None:
                stop_status = "still_alive"
                stop_date = "1/1/1800"
                continue

            if is_red and stop_date is None:
                stop_date = date
                stop_status = "completed"
                break

        if intro_date is not None and stop_date is not None:
            sheet_person_num += 1
            entity_id = f"{prefix}{sheet_person_num}"

            duration_days = (stop_date - intro_date).days
            yearfrac = calculate_yearfrac(intro_date, stop_date)

            records.append({
                "sheet": sheet_name,
                "entity_id": entity_id,
                "start_date": intro_date,
                "stop_date": stop_date,
                "duration_days": duration_days,
                "yearfrac": yearfrac,
                "status": stop_status,
                "red_census": "Yes" if had_red_census else "No",
                "green_census": "Yes" if had_green_census else "No"
            })


expected_columns = ["sheet", "entity_id", "start_date", "stop_date", "duration_days", "yearfrac", "status", "red_census", "green_census"]

df = pd.DataFrame(records, columns=expected_columns)

if not df.empty:
    df = df.sort_values(
        by=["sheet", "entity_id", "start_date"]
    ).reset_index(drop=True)
else:
    df = df.reset_index(drop=True)

print(df)


df.to_excel(OUTPUT_FILE, index=False)

print(f"\nFinished! Output saved to {OUTPUT_FILE}")